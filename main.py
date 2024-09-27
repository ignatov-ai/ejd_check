# -*- coding: utf8 -*-
import shutil

from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Alignment
from openpyxl.utils import get_column_letter

from datetime import datetime
import os
import sys
import warnings

def find_data(sheet, search_value):
    found = False
    for row in sheet.iter_rows():
        for cell in row:
            if cell.value == search_value:
                coordinate = cell.coordinate
                column = ""
                column_number = 0
                row = ""
                for char in coordinate:
                    if char.isalpha():  # Проверяем, является ли символ буквой
                        column += char
                    elif char.isdigit():  # Проверяем, является ли символ цифрой
                        row += char

                for i, letter in enumerate(reversed(column)):
                    column_number += (ord(letter.upper()) - ord('A') + 1) * (26 ** i)
                return column, column_number, row

    if not found:
        return None, None, None

warnings.simplefilter("ignore")

data_str = ['1', '2', '3', '4', '5']
marks_str = ['2', '3', '4', '5']
marks_int = [2, 3, 4, 5]

start_time = datetime.now()

# параллели для проверки
grade = "9-11"
# папка с журналами
journals_folder = "journals_180924\\" + grade
# journals_folder = "journals\\"

#######################################################
#################### Загрузка УП ######################
#######################################################
# up_file = 'УП'
# up_book = load_workbook("up\\" + up_file + '.xlsx')
#
# uchebniy_plan = []
#
# up_sheet = up_book.active
#
# for row in up_sheet.iter_rows(values_only=True):
#     uchebniy_plan.append(list(row))
#print(uchebniy_plan)

#######################################################
#######################################################
#######################################################

# Получаем текущую директорию
current_dir = os.getcwd()

# Создаем список для хранения имен файлов
file_list = []

# Перебираем файлы в текущей директории
# for root, dirs, files in os.walk(current_dir + '\\journals'):
for root, dirs, files in os.walk(current_dir + '\\' + journals_folder):
    for filename in files:
        if filename[-5:] == '.xlsx':
            file_list.append(filename)
# Выводим список имен файлов
# print(file_list)

# книга замечаний по ОТМЕТКАМ
comment_book = Workbook()
comment_book_sheet = comment_book.active
comment_book_sheet.append(['Класс',
                           'Ученик',
                           'Предмет',
                           'Минимум отметок',
                           'Кол-во отметок',
                           'Не хватает отметок'])
comment_book_sheet.column_dimensions['A'].width = 6
comment_book_sheet.column_dimensions['B'].width = 50
comment_book_sheet.column_dimensions['C'].width = 50
comment_book_sheet.column_dimensions['D'].width = 5
comment_book_sheet.column_dimensions['E'].width = 5
comment_book_sheet.column_dimensions['F'].width = 5

first_row = comment_book_sheet["1"]
for cell in first_row:
    style = Alignment(text_rotation=90, horizontal='center', vertical='center')
    cell.alignment = style

# книга замечаний по КТП и ДЗ
    comment_book_dz_ktp = Workbook()
    comment_book_sheet_dz_ktp = comment_book_dz_ktp.active
    comment_book_sheet_dz_ktp.append(['Класс',
                                      'Предмет',
                                      'Группа',
                                      'Кол-во уроков',
                                      'Кол-во уроков без КТП',
                                      'Кол-во уроков без ДЗ'])

    comment_book_sheet_dz_ktp.column_dimensions['A'].width = 6
    comment_book_sheet_dz_ktp.column_dimensions['B'].width = 50
    comment_book_sheet_dz_ktp.column_dimensions['C'].width = 50
    comment_book_sheet_dz_ktp.column_dimensions['D'].width = 5
    comment_book_sheet_dz_ktp.column_dimensions['E'].width = 5
    comment_book_sheet_dz_ktp.column_dimensions['F'].width = 5

    first_row = comment_book_sheet_dz_ktp["1"]
    for cell in first_row:
        style = Alignment(text_rotation=90, horizontal='center', vertical='center')
        cell.alignment = style

# выбирается файл для обработки
for file in file_list:
    file = file.split('.xlsx')[0]

    class_name = file.split(";")[0]
    lesson = file.split(";")[1]
    group = file.split(";")[2]

    print(class_name, lesson, group)

    #######################################################################################
    ####################  Проверка журналов на выставление отметок  #######################
    #######################################################################################

    # книга для проверки отметок
    # book = load_workbook("journals\\" + file + '.xlsx')
    book = load_workbook(journals_folder + "\\" + file + '.xlsx')
    sheet = book.active

    # удаляем столбцы с ДЗ
    sheet.delete_cols(20, 23)

       # получаем список объединенных диапазонов и разъединяем объединенные ячейки в книге ОТМЕТОК
    merged_cells = list(map(str, sheet.merged_cells.ranges))

    for item in merged_cells:
        try:
            sheet.unmerge_cells(item)
        except KeyError:
            continue

    # перемещаем куски оценок в одну общую часть
    for part in range(1, 6):
        m_range = 'C'+str(1+part*50)+':S'+str((part+1)*50)
        sheet.move_range(m_range, rows=-50*part, cols=17*part)

    # считаем количество учеников на листе
    students_count = 1

    if sheet['A' + str(students_count)].value == "Нет оценок":
        continue

    while sheet['A'+str(students_count)].value != '':
        # print(class_name, lesson, "(", sheet['A' + str(students_count)].value, ")", students_count)
        students_count += 1

    #students_count -= 3
    # print(students_count)

    # удаляем пустые строки
    sheet.delete_rows(students_count, sheet.max_row)

    # заменяем ширину всех столбцов с оценками
    for i in range(3, int(sheet.max_column)):
        sheet.column_dimensions[get_column_letter(i)].width = 2

    # меняем тип ячейки на число, если это возможно
    for row in range(students_count-3):
        for col in range(3, int(sheet.max_column)):
            ch_range = get_column_letter(col) + str(row+3)
            if sheet[ch_range].value in data_str:
                sheet[ch_range].value = int(sheet[ch_range].value)

    # удаление лишних столбцов
    delete_columns_list = ['Т1', 'Т2', 'Т3', 'М1', 'М2', 'М3', 'М4', 'М5', 'М6', 'П1', 'П2', 'Г', 'Э', 'А']
    for s in delete_columns_list:
        del_column, del_column_number, del_row = find_data(sheet, s)
        if (del_column, del_column_number, del_row) != (None, None, None):
            sheet.delete_cols(del_column_number)

    #подсчет проведенных уроков на данную дату и удаление не нужных столбцов
    lessons_count = 0
    month_count = 1
    old_day_value = 0
    curent_month = 9
    curent_day = 0
    last_column_index = 1

    now_month = datetime.now().month
    now_day = datetime.now().day
    # print(now_month, now_day)


    for i in range(2, sheet.max_column+1):
        new_cell_value = sheet[str(get_column_letter(i)) + "2"].value
        # print(new_cell_value)

        if new_cell_value != "" and new_cell_value is not None:
            curent_day = int(new_cell_value)
            lessons_count += 1
        elif new_cell_value == "":
            break

        if curent_day < int(old_day_value):
            month_count += 1
            curent_month += 1

        # print('месяц:', curent_month, 'день', curent_day)

        if curent_month >= now_month and curent_day >= now_day:
            #print("дальше еще уроков не было")
            last_column_index = i + 2 + 1
            break

        old_day_value = curent_day

    # print(lessons_count)
    # print(month_count)

    # sheet.delete_cols(last_column_index, sheet.max_column)

    # сохранение переработанной книги в data
    # book.save('data\\' + file + '_marks.xlsx')
    book.close()

    # МИНИМАЛЬНОЕ КОЛИЧЕСТВО ОТМЕТОК НА ДАННЫЙ МОМЕНТ
    min_marks_for_now = 1

    # подсчет количества оценок по ученикам
    for row in range(4, sheet.max_row+1):
        marks_count_for_student = 0
        current_student = sheet['B' + str(row)].value[:-2]
        print(class_name, lesson, current_student, end=": ")
        for col in range(3, int(sheet.max_column+2)):
            if sheet[get_column_letter(col) + str(row)].value in marks_int:
                marks_count_for_student += 1
        # print(marks_count_for_student)

        # добавляем замечание о недостаточном количестве отметок у учащегося
        if marks_count_for_student < min_marks_for_now:
            comment_book_sheet.append([class_name,
                                       current_student,
                                       lesson,
                                       min_marks_for_now,
                                       marks_count_for_student,
                                       min_marks_for_now - marks_count_for_student])



    comment_book.save('comments\ВСЕ ЗАМЕЧАНИЯ ПО ПРОВЕРКЕ НАКОПЛЯЕМОСТИ ОЦЕНОК ' + grade + '.xlsx')

    #######################################################################################
    #########################  Прокерка журналов на КТП И ДЗ  #############################
    #######################################################################################

    # книга для проверки ДЗ и КТП
    book_dz = load_workbook(journals_folder + "\\" + file + '.xlsx')
    sheet_dz = book_dz.active

    # получаем список объединенных диапазонов и разъединяем объединенные ячейки в книге ДЗ и КТП
    merged_cells = list(map(str, sheet_dz.merged_cells.ranges))

    for item in merged_cells:
        try:
            sheet_dz.unmerge_cells(item)
        except KeyError:
            continue

    # удаляем столбцы с отметками
    sheet_dz.delete_cols(1, 20)
    sheet_dz.delete_cols(4, 25)

    # удаляем пустые строки delete_rows(первая строка, количество строк после)
    for row in reversed(range(1, sheet_dz.max_row)):
        sum = 0
        for col in range(1, 4):
            cell = sheet_dz[get_column_letter(col) + str(row)].value
            if cell is None or cell == "":
                sum += 1
        if sum == 3:
            sheet_dz.delete_rows(row, 1)

    # удаляем ЛИШНИЕ СТРОКИ С ДОП ИНФОРМАЦИЕЙ delete_rows(первая строка, количество строк после)
    for row in reversed(range(1, sheet_dz.max_row)):
        cell = sheet_dz['A' + str(row)].value
        if len(cell) > 5 or cell == 'Дата':
            sheet_dz.delete_rows(row, 1)

    # форматируем книгу с ДЗ и КТП
    sheet_dz.column_dimensions['A'].width = 10
    sheet_dz.column_dimensions['B'].width = 70
    sheet_dz.column_dimensions['C'].width = 70

    # считаем количество уроков "БЕЗ ТЕМЫ"
    wo_ktp_count = 0
    wo_dz_count = 0
    for row in range(1, sheet_dz.max_row):
        cell_ktp = sheet_dz['B' + str(row)].value
        if cell_ktp == 'Без темы':
            wo_ktp_count += 1

        cell_dz = sheet_dz['C' + str(row)].value
        if cell_dz == 'не задано':
            wo_dz_count += 1
    if wo_ktp_count > 0:
        comment_book_sheet_dz_ktp.append([  class_name,
                                            lesson,
                                            group,
                                            sheet_dz.max_row-1,
                                            wo_ktp_count,
                                            wo_dz_count])

        # print(class_name, lesson, group, 'нет тем уроков:', wo_ktp_count, 'из:', sheet_dz.max_row-1,
        #                                   '| Задано ДЗ:', sheet_dz.max_row-1-wo_dz_count, 'из:', sheet_dz.max_row-1)

    # book_dz.save('data\\' + file + '_dz_ktp.xlsx')
    book_dz.close()
    comment_book_dz_ktp.save('comments\ВСЕ ЗАМЕЧАНИЯ ПО ПРОВЕРКЕ ДЗ И КТП ' + grade + '.xlsx')

    # перемещаем обработанный файл в папку DONE
    # source_file = 'journals\\' + file + '.xlsx'
    # destination_folder = 'done\\' + file + '.xlsx'
    # shutil.move(source_file, destination_folder)
end_time = datetime.now()

print(start_time, end_time)

comment_book.close()
comment_book_dz_ktp.close()