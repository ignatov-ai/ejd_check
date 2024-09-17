##########################################################################
### ПРОГРАММА ДЛЯ ПРОВЕРКИ ВЫСТАВЛЕННЫХ ОЦЕНОК ПО РАСШИРЕННОЙ ВЫГРУЗКЕ ###
##########################################################################

import openpyxl.styles
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Alignment
from openpyxl.utils import get_column_letter

import datetime
import os
import warnings

warnings.simplefilter("ignore")

marks_str = ['0', '1', '2', '3', '4', '5']
marks_int = [0, 1, 2, 3, 4, 5]

# Получаем текущую директорию
current_dir = os.getcwd()

# Создаем список для хранения имен файлов
file_list = []

# Перебираем файлы в текущей директории
for root, dirs, files in os.walk(current_dir + '\\data'):
    for filename in files:
        if filename[-5:] == '.xlsx':
            file_list.append(filename)
print(file_list)

# общая книга замечаний по накопляемости
comment_book = openpyxl.Workbook()
comment_book_sheet = comment_book.active
comment_book_sheet.append(['Класс', 'Предмет', 'Ученик', 'Минимум оценок',
                           'Кол-во оценок', 'Не хватает оценок'])
comment_book_sheet.column_dimensions['A'].width = 6
comment_book_sheet.column_dimensions['B'].width = 50
comment_book_sheet.column_dimensions['C'].width = 50
comment_book_sheet.column_dimensions['D'].width = 5
comment_book_sheet.column_dimensions['E'].width = 5
comment_book_sheet.column_dimensions['F'].width = 5

first_row = comment_book_sheet["1"]
for cell in first_row:
    style = Alignment(text_rotation=90, horizontal='center', vertical='center')
    cell.alignment = style

# общая книга замечаний по выставлению оценок за аттестационный период
comment_avg_book = openpyxl.Workbook()
comment_avg_book_sheet = comment_avg_book.active
comment_avg_book_sheet.append(['Класс', 'Предмет', 'Ученик', 'Выставленная оценка',
                           'Требуемая оценка'])
comment_avg_book_sheet.column_dimensions['A'].width = 6
comment_avg_book_sheet.column_dimensions['B'].width = 50
comment_avg_book_sheet.column_dimensions['C'].width = 50
comment_avg_book_sheet.column_dimensions['D'].width = 6
comment_avg_book_sheet.column_dimensions['E'].width = 6

first_row = comment_avg_book_sheet["1"]
for cell in first_row:
    style = Alignment(text_rotation=90, horizontal='center', vertical='center')
    cell.alignment = style

# выбирается файл для обработки
for file in file_list:
    file = file.split('.xlsx')[0]

    book = load_workbook("data\\" + file + '.xlsx')

    comment_class_book = openpyxl.Workbook()
    comment_class_book_sheet = comment_class_book.active
    comment_class_book_sheet.append(['Класс', 'Предмет', 'Ученик', 'Средний балл', 'Выставлено',
                               'Кол-во оценок', 'Не хватает оценок'])
    comment_class_book_sheet.column_dimensions['A'].width = 6
    comment_class_book_sheet.column_dimensions['B'].width = 50
    comment_class_book_sheet.column_dimensions['C'].width = 50
    comment_class_book_sheet.column_dimensions['D'].width = 5
    comment_class_book_sheet.column_dimensions['E'].width = 5
    comment_class_book_sheet.column_dimensions['F'].width = 5
    comment_class_book_sheet.column_dimensions['G'].width = 5

    first_row = comment_class_book_sheet["1"]
    for cell in first_row:
        style = Alignment(text_rotation=90, horizontal='center', vertical='center')
        cell.alignment = style

    # получаем список листов
    tabs = book.sheetnames

    # выбирается лист для работы
    for sheet_name in tabs:
        date = datetime.datetime.now()

        sheets = book[sheet_name]

        # определяем предмет
        lesson = sheets['U41'].value.split(',')[-1].lstrip().rstrip()

        # удаляем столбцы с ДЗ
        sheets.delete_cols(20, 23)

        # получаем список объединенных диапазонов и разъединяем объединенные ячейки
        merged_cells = list(map(str, sheets.merged_cells.ranges))

        for item in merged_cells:
            try:
                sheets.unmerge_cells(item)
            except KeyError:
                continue

        # перемещаем куски оценок в одну общую часть
        parts = 0
        while sheets['A'+str(parts*50 + 1)].value == '№':
            parts += 1

        for part in range(1, parts+1):
            m_range = 'C'+str(1+part*50)+':S'+str((part+1)*50)
            sheets.move_range(m_range, rows=-50*part, cols=17*part)

        # добавляем колонку КОЛИЧЕСТВО ОЦЕНОК и КОЛИЧЕСТВО ПРОПУСКОВ и подсчитываем количество оценок и пропусков по каждому учащемуся
        sheets[get_column_letter(parts * 17 + 3) + '2'].value = 'Кол-во оц.'
        sheets.column_dimensions[get_column_letter(parts * 17 + 3)].width = 11

        sheets[get_column_letter(parts * 17 + 4) + '2'].value = 'Ср. балл'
        sheets.column_dimensions[get_column_letter(parts * 17 + 3)].width = 11

        # считаем количество уроков за период
        marks_for_period = 0
        n = 3

        while sheets[get_column_letter(n) + str(2)].value != 'АП1':
            #print(n, get_column_letter(n) + str(3), sheets[get_column_letter(n) + str(3)].value)
            if sheets[get_column_letter(n) + str(3)].value == 'оц':
                marks_for_period += 1
            n += 1

        if marks_for_period <= 34:
            up_marks_count = 3
        elif 34 < marks_for_period <= 68:
            up_marks_count = 5
        elif marks_for_period > 68:
            up_marks_count = 7
        print('Проведено занятий:', marks_for_period, 'Нужно оценок:', up_marks_count)

        # считаем количество учеников на листе
        students_count = 1
        while sheets['A'+str(students_count)].value != '':
            students_count += 1
        #students_count -= 3

        # удаляем пустые строки
        sheets.delete_rows(students_count, sheets.max_row)

        # меняем тип ячейки на число, если это возможно
        for row in range(students_count-4):
            marks_count = 0
            marks = []
            for col in range(3, parts*17+3):
                mark_cell = get_column_letter(col) + str(row+4)
                koeff_cell = get_column_letter(col+1) + str(row+4)

                if sheets[get_column_letter(col) + '2'].value == 'АП1':
                    if sheets[get_column_letter(col) + '2'].value in marks_str:
                        sheets[mark_cell].value = int(sheets[mark_cell].value)
                    ap_mark = sheets[mark_cell].value

                    # проверка что оценка вообще стоит
                    if ap_mark == '':
                        #print(sheets['B' + str(row + 4)].value, ap_mark)
                        # добавляем замечние в файл КЛАССА
                        comment_avg_book_sheet.append(
                            [file.split()[0], lesson, sheets['B' + str(row + 4)].value, 'НЕТ ОЦЕНКИ'])


                    break

                if sheets[mark_cell].value in marks_str and sheets[get_column_letter(col) + str(3)].value == 'оц':
                    sheets[mark_cell].value = int(sheets[mark_cell].value)

                    marks_count += 1

                    if sheets[koeff_cell].value == '1':
                        marks.append(sheets[mark_cell].value)
                    elif sheets[koeff_cell].value == '2':
                        marks.append(sheets[mark_cell].value)
                        marks.append(sheets[mark_cell].value)
                    elif sheets[koeff_cell].value == '3':
                        marks.append(sheets[mark_cell].value)
                        marks.append(sheets[mark_cell].value)
                        marks.append(sheets[mark_cell].value)

            fill_cell = sheets[get_column_letter(parts * 17 + 4) + str(row + 4)]

            if marks_count >= up_marks_count:
                avg_marks = sum(marks)/len(marks)

                if avg_marks < 2.6:
                    avg_marks_round = '2'
                elif avg_marks < 3.6:
                    avg_marks_round = '3'
                elif avg_marks < 4.6:
                    avg_marks_round = '4'
                else:
                    avg_marks_round = '5'

                if ap_mark == avg_marks_round:
                    fill_cell.fill = PatternFill(fill_type='solid', fgColor="85EB6A")
                else:
                    fill_cell.fill = PatternFill(fill_type='solid', fgColor="FA7080")

                    # добавляем замечние в файл КЛАССА
                    comment_class_book_sheet.append(
                        [file.split()[0], lesson, sheets['B' + str(row + 4)].value, avg_marks, ap_mark, marks_count, up_marks_count-marks_count])

                    # добавляем замечние в ОБЩУЮ СВОДКУ
                    comment_book_sheet.append(
                        [file.split()[0], lesson, sheets['B' + str(row + 4)].value, ap_mark, marks_count, up_marks_count-marks_count])

                    # добавляем замечние в файл КЛАССА
                    comment_avg_book_sheet.append(
                        [file.split()[0], lesson, sheets['B' + str(row + 4)].value, ap_mark, int(avg_marks_round)])

            else:
                avg_marks = 'А/З'

                if ap_mark == avg_marks:
                    fill_cell.fill = PatternFill(fill_type='solid', fgColor="85EB6A")
                else:
                    fill_cell.fill = PatternFill(fill_type='solid', fgColor="FA7080")

            sheets[get_column_letter(parts*17+3)+str(row+4)].value = marks_count
            sheets[get_column_letter(parts * 17 + 4) + str(row + 4)].value = avg_marks

        # зменяем ширину всех столбцов с оценками
        for i in range(3, parts * 17 + 3):
            sheets.column_dimensions[get_column_letter(i)].width = 2

        # Вставляем класс, учителя и дату
        sheets.insert_rows(1, 3)
        sheets['B1'] = file
        sheets['B2'] = lesson

        print(file, ':', lesson, '- Проверен (', date, ')')

    book.save('checked\\' + file + ' ПРОВЕРЕНО.xlsx')
    book.close()

    # перемещаем обработанный файл в папку DONE
    # source_file = 'data\\' + file + '.xlsx'
    # destination_folder = 'done\\' + file + '.xlsx'
    # shutil.move(source_file, destination_folder)

    comment_class_book.save('comments\\' + file + ' ЗАМЕЧАНИЯ.xlsx')
    comment_class_book.close()

    comment_book.save('comments\ВСЕ ЗАМЕЧАНИЯ ПО ПРОВЕРКЕ НАКОПЛЯЕМОСТИ ОЦЕНОК.xlsx')
    comment_book.close()

    comment_avg_book.save('avg_comments\ВСЕ ЗАМЕЧАНИЯ ПО ПРОВЕРКЕ ВЫСТАВЛЕННЫХ ОЦЕНОК.xlsx')
    comment_avg_book.close()