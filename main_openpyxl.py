from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Alignment
from openpyxl.utils import get_column_letter

import datetime
import os
import sys
import warnings

#print('Hello ворлд')

warnings.simplefilter("ignore")

marks_str = ['2', '3', '4', '5']
marks_int = [2, 3, 4, 5]

#######################################################
#################### Загрузка УП ######################
#######################################################
up_file = 'УП'
up_book = load_workbook("up\\" + up_file + '.xlsx')

uchebniy_plan = []

up_sheet = up_book.active

for row in up_sheet.iter_rows(values_only=True):
    uchebniy_plan.append(list(row))

#print(uchebniy_plan)

#######################################################
#######################################################
#######################################################

# Получаем текущую директорию
current_dir = os.getcwd()

# Создаем список для хранения имен файлов
file_list = []

# Перебираем файлы в текущей директории
for root, dirs, files in os.walk(current_dir + '\data'):
    for filename in files:
        if filename[-5:] == '.xlsx':
            file_list.append(filename)
# Выводим список имен файлов
print(file_list)

# общая книга замечаний
comment_book = Workbook()
comment_book_sheet = comment_book.active
#comment_book_sheet.append(['Класс', 'Предмет', 'Учитель', 'Ученик', 'Минимум оценок', - убрал столбик Учитель
comment_book_sheet.append(['Класс', 'Ученик', 'Предмет', 'Минимум оценок',
                           'Кол-во оценок', 'Не хватает оценок'])
comment_book_sheet.column_dimensions['A'].width = 6
comment_book_sheet.column_dimensions['B'].width = 50
comment_book_sheet.column_dimensions['C'].width = 50
comment_book_sheet.column_dimensions['D'].width = 5
comment_book_sheet.column_dimensions['E'].width = 5
comment_book_sheet.column_dimensions['F'].width = 5
#comment_book_sheet.column_dimensions['G'].width = 5

first_row = comment_book_sheet["1"]
for cell in first_row:
    style = Alignment(text_rotation=90, horizontal='center', vertical='center')
    cell.alignment = style

# выбирается файл для обработки
for file in file_list:
    file = file.split('.xlsx')[0]

    book = load_workbook("data\\" + file + '.xlsx')

    comment_class_book = openpyxl.Workbook()
    comment_class_book_sheet = comment_class_book.active
    #comment_class_book_sheet.append(['Класс', 'Предмет', 'Учитель', 'Ученик', 'Минимум оценок', - убрал столбик Учитель
    comment_class_book_sheet.append(['Класс', 'Ученик', 'Предмет', 'Минимум оценок',
                               'Кол-во оценок', 'Не хватает оценок'])
    comment_class_book_sheet.column_dimensions['A'].width = 6
    comment_class_book_sheet.column_dimensions['B'].width = 50
    comment_class_book_sheet.column_dimensions['C'].width = 50
    comment_class_book_sheet.column_dimensions['D'].width = 5
    comment_class_book_sheet.column_dimensions['E'].width = 5
    comment_class_book_sheet.column_dimensions['F'].width = 5
    #comment_class_book_sheet.column_dimensions['G'].width = 5

    first_row = comment_class_book_sheet["1"]
    for cell in first_row:
        style = Alignment(text_rotation=90, horizontal='center', vertical='center')
        cell.alignment = style

    # получаем список листов
    tabs = book.sheetnames

    # выбирается лист для работы
    for sheet_name in tabs:
        #print(file, sheet_name)
        date = datetime.datetime.now()

        sheets = book[sheet_name]

        # определяем предмет
        lesson = sheets['U41'].value.split(',')[-1].lstrip().rstrip()

        # определяем учителя
        teacher_split = sheets['U43'].value.split()[1:4]
        teacher = ' '.join(teacher_split).lstrip().rstrip()

        # удаляем столбцы с ДЗ
        sheets.delete_cols(20, 23)

        # получаем список объединенных диапазонов и разъединяем объединенные ячейки
        merged_cells = list(map(str, sheets.merged_cells.ranges))

        for item in merged_cells:
            try:
                sheets.unmerge_cells(item)
            except KeyError:
                continue

        # перемещаем куски оценок в одну общую часть
        for part in range(1, 6):
            m_range = 'C'+str(1+part*50)+':S'+str((part+1)*50)
            sheets.move_range(m_range, rows=-50*part, cols=17*part)

        # считаем количество учеников на листе
        students_count = 1
        while sheets['A'+str(students_count)].value != '':
            #print(sheets['A' + str(students_count)].value,students_count)
            students_count += 1
        #students_count -= 3

        # удаляем пустые строки
        sheets.delete_rows(students_count, sheets.max_row)

        # зменяем ширину всех столбцов с оценками
        for i in range(3, int(sheets.max_column)):
            sheets.column_dimensions[get_column_letter(i)].width = 2

        # меняем тип ячейки на число, если это возможно
        for row in range(students_count-3):
            for col in range(3, int(sheets.max_column)):
                ch_range = get_column_letter(col) + str(row+3)
                if sheets[ch_range].value in marks_str:
                    sheets[ch_range].value = int(sheets[ch_range].value)

        # добавляем колонку КОЛИЧЕСТВО ОЦЕНОК и КОЛИЧЕСТВО ПРОПУСКОВ и подсчитываем количество оценок и пропусков по каждому учащемуся
        sheets['CJ2'].value = 'Количество оценок'
        sheets['CK2'].value = 'Количество пропусков'

        i = 3
        for col in sheets.iter_rows(min_col=3, max_col=87, min_row=3, max_row=students_count-1):
            marks_count = 0
            n_count = 0
            for cell in col:
                if cell.value in marks_int:
                    marks_count += 1
                elif cell.value == 'н':
                    n_count += 1

            sheets['CJ'+str(i)].value = marks_count
            sheets['CK' + str(i)].value = n_count
            i += 1

        # Вставляем класс, учителя и предмет
        sheets.insert_rows(1, 4)
        sheets['B1'] = file
        sheets['B2'] = lesson
        sheets['B3'] = teacher

        #######################################################
        ###########  Сравнение кол-ва уроков с УП  ############
        #######################################################
        up_class_index = uchebniy_plan[0].index(file.lower())
        #print(up_class_index, file.lower())

        up_lesson_index = 0
        for i in range (len(uchebniy_plan)):
            if uchebniy_plan[i][0] == lesson.lower():
                up_lesson_index = i
                #print(up_lesson_index, lesson.lower(),uchebniy_plan[up_lesson_index][up_class_index])
                break
        if up_lesson_index == 0:
            print('ПРОВЕРЬТЕ НАЗВАНИЕ ПРЕДМЕТА В УЧЕБНОМ ПЛАНЕ!!!\n'
                  'НАЗВАНИЕ ДОЛЖНО СОВПАДАТЬ С ВЫГРУЗКОЙ ИЗ ЭЖД')
            sys.exit()

        up_lesson_nagruzka = uchebniy_plan[up_lesson_index][up_class_index]

        #print(up_lesson_nagruzka)

        up_marks_number = up_lesson_nagruzka * 2 + 1
        if up_marks_number > 7:
            up_marks_number = 7

        #print(file.lower(), lesson.lower(), up_lesson_nagruzka, up_marks_number)

        for row in range(students_count-3):
            ch_cell = 'CJ' + str(row+7)
            fill_cell = sheets[ch_cell]
            #print(ch_cell, sheets[ch_cell].value)

            student_fio = sheets['B' + str(row + 7)].value

            #print(student_fio, student_fio[-5:])
            if student_fio[-5:] != '.2023':
                if fill_cell.value >= up_marks_number:
                    fill_cell.fill = PatternFill(fill_type='solid', fgColor="85EB6A")
                else:
                    fill_cell.fill = PatternFill(fill_type='solid', fgColor="FA7080")
                    # добавляем замечние в файл КЛАССА
                    #comment_class_book_sheet.append([file, lesson, teacher, sheets['B' + str(row + 7)].value, up_marks_number,  - убрал столбик Учитель
                    comment_class_book_sheet.append([file, sheets['B' + str(row + 7)].value, lesson, up_marks_number,
                                                     fill_cell.value, up_marks_number - fill_cell.value])

                    if up_marks_number-fill_cell.value == 1:
                        comment_class_book_cell = comment_class_book_sheet['F' + str(len(comment_class_book_sheet['A']))]
                        comment_class_book_cell.fill = PatternFill(fill_type='solid', fgColor="FFE697")
                    elif up_marks_number - fill_cell.value == 2:
                        comment_class_book_cell = comment_class_book_sheet['F' + str(len(comment_class_book_sheet['A']))]
                        comment_class_book_cell.fill = PatternFill(fill_type='solid', fgColor="FFD040")
                    elif up_marks_number - fill_cell.value == 3:
                        comment_class_book_cell = comment_class_book_sheet['F' + str(len(comment_class_book_sheet['A']))]
                        comment_class_book_cell.fill = PatternFill(fill_type='solid', fgColor="F9A13E")
                    elif up_marks_number - fill_cell.value > 3:
                        comment_class_book_cell = comment_class_book_sheet['F' + str(len(comment_class_book_sheet['A']))]
                        comment_class_book_cell.fill = PatternFill(fill_type='solid', fgColor="F36149")

                    # добавляем замечние в СВОДНЫЙ ОТЧЕТ
                    comment_book_sheet.append(
                    #    [file, lesson, teacher, sheets['B' + str(row + 7)].value, up_marks_number,  - убрал столбик Учитель
                        [file, sheets['B' + str(row + 7)].value, lesson, up_marks_number,
                         fill_cell.value, up_marks_number - fill_cell.value])

                    if up_marks_number-fill_cell.value == 1:
                        comment_book_cell = comment_book_sheet['F' + str(len(comment_book_sheet['A']))]
                        comment_book_cell.fill = PatternFill(fill_type='solid', fgColor="FFE697")
                    elif up_marks_number - fill_cell.value == 2:
                        comment_book_cell = comment_book_sheet['F' + str(len(comment_book_sheet['A']))]
                        comment_book_cell.fill = PatternFill(fill_type='solid', fgColor="FFD040")
                    elif up_marks_number - fill_cell.value == 3:
                        comment_book_cell = comment_book_sheet['F' + str(len(comment_book_sheet['A']))]
                        comment_book_cell.fill = PatternFill(fill_type='solid', fgColor="F9A13E")
                    elif up_marks_number - fill_cell.value > 3:
                        comment_book_cell = comment_book_sheet['F' + str(len(comment_book_sheet['A']))]
                        comment_book_cell.fill = PatternFill(fill_type='solid', fgColor="F36149")

        #######################################################
        #######################################################
        #######################################################

        # ставим дату и время проверки в ячейку B2
        sheets['B4'] = date

        # удаление пустых столбцов перед количеством оценок и пропусков
        i = 87
        flag = False
        while True:
            # Проверка, что столбец не содержит данных
            column_a = get_column_letter(i)

            for j in range (1,50):
                if sheets[column_a + str(j)].value not in ['', None]:
                    #print('Столбец ', column_a,' содержит данные')
                    flag = True
                    break
            if flag:
                break
            else:
                i -= 1

        #print('Смещение с', 87, 'на', 87-i)

        m_range = 'CJ1:CK50'
        sheets.move_range(m_range, rows=0, cols=-(87-i))

        sheets.column_dimensions[get_column_letter(i + 1)].width = 11
        sheets.column_dimensions[get_column_letter(i + 2)].width = 11

        for row in sheets.iter_rows():
            for cell in row:
                cell.border = None
                cell.font = None

        print(file, ':', lesson, '- Проверен (', date, ')')

    book.save('checked\\' + file + ' ПРОВЕРЕНО.xlsx')
    book.close()

    # перемещаем обработанный файл в папку DONE
    # source_file = 'data\\' + file + '.xlsx'
    # destination_folder = 'done\\' + file + '.xlsx'
    # shutil.move(source_file, destination_folder)

    comment_class_book.save('comments\\' + file + ' ЗАМЕЧАНИЯ.xlsx')
    comment_class_book.close()

    comment_book.save('comments\ВСЕ ЗАМЕЧАНИЯ ПО ПРОВЕРКЕ НАКОПЛЯЕМОСТИ ОЦЕНОК.xlsx')
    comment_book.close()