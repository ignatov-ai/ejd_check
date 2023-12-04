import openpyxl.styles
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Alignment
from openpyxl.utils import get_column_letter

import datetime
import os
import sys
import warnings

warnings.simplefilter("ignore")

marks_str = ['2', '3', '4', '5']
marks_int = [2, 3, 4, 5]

# Получаем текущую директорию
current_dir = os.getcwd()

# Создаем список для хранения имен файлов
file_list = []

#######################################################
#################### Загрузка УП ######################
#######################################################
up_file = 'УП'
up_book = load_workbook("up\\" + up_file + '.xlsx')

uchebniy_plan = []

up_sheet = up_book.active

for row in up_sheet.iter_rows(values_only=True):
    uchebniy_plan.append(list(row))

# Перебираем файлы в текущей директории
for root, dirs, files in os.walk(current_dir + '\\class_data'):
    for filename in files:
        if filename[-5:] == '.xlsx':
            file_list.append(filename)
# Выводим список имен файлов
print(file_list)

for file in file_list:
    file = file.split('.xlsx')[0]

    avg_book = load_workbook("avg\Отчёт по средним баллам " + file + ' класс..xlsx')
    avg_book_sheet = avg_book.active

    avg_class_data = []

    for row in range(7, avg_book_sheet.max_row+1):
        avg_row_class_data = [avg_book_sheet['A' + str(row)].value]
        for col in range(2, int(avg_book_sheet.max_column)+1):
            avg_book_sheet[get_column_letter(col) + str(row)].value = float(avg_book_sheet[get_column_letter(col) + str(row)].value)
            if avg_book_sheet[get_column_letter(col) + str(row)].value < 2.6:
                avg_book_sheet[get_column_letter(col) + str(row)].value = 'А/З'
            elif 2.6 <= avg_book_sheet[get_column_letter(col) + str(row)].value < 3.6:
                avg_book_sheet[get_column_letter(col) + str(row)].value = 3
            elif 3.6 <= avg_book_sheet[get_column_letter(col) + str(row)].value < 4.6:
                avg_book_sheet[get_column_letter(col) + str(row)].value = 4
            else:
                avg_book_sheet[get_column_letter(col) + str(row)].value = 5

            avg_row_class_data.append(avg_book_sheet[get_column_letter(col) + str(row)].value)
        avg_class_data.append(avg_row_class_data)

    #print(avg_class_data)

    avg_book_sheet.row_dimensions[6].height = 350
    rotate_row = avg_book_sheet["6"]
    for cell in rotate_row:
        style = Alignment(text_rotation=90, horizontal='center', vertical='center')
        cell.alignment = style
        cell.value = cell.value.lower()

    for col in range(2, int(avg_book_sheet.max_column) + 1):
        avg_book_sheet.column_dimensions[get_column_letter(col)].width = 4

    # составляем список фамилий и имен с листа средних баллов
    students_indexes = []
    for i in range(len(avg_class_data)):
        students_indexes.append(' '.join(avg_class_data[i][0].split()[:2]))

    #print(students_indexes)

    # составляем список предметов с листа средних баллов
    lessons_indexes = []
    for i in range(2, len(avg_class_data[6])+1):
        lessons_indexes.append(avg_book_sheet[get_column_letter(i)+'6'].value)

    #print(lessons_indexes)

    #выбирается файл для обработки
    book = load_workbook('class_data\\' + file + '.xlsx')

    # получаем список листов
    tabs = book.sheetnames

    # выбирается лист для работы
    for sheet_name in tabs:

        sheets = book[sheet_name]

        # определяем предмет
        lesson = sheets['U41'].value.split(',')[-1].lstrip().rstrip().lower()

        # удаляем столбцы с ДЗ
        sheets.delete_cols(20, 23)

        # получаем список объединенных диапазонов и разъединяем объединенные ячейки
        merged_cells = list(map(str, sheets.merged_cells.ranges))

        for item in merged_cells:
            try:
                sheets.unmerge_cells(item)
            except KeyError:
                continue

        # перемещаем куски оценок в одну общую часть
        for part in range(1, 6):
            m_range = 'C'+str(1+part*50)+':S'+str((part+1)*50)
            sheets.move_range(m_range, rows=-50*part, cols=17*part)

        # считаем количество учеников на листе
        students_count = 1
        while sheets['A'+str(students_count)].value != '':
            students_count += 1

        # удаляем пустые строки
        sheets.delete_rows(students_count, sheets.max_row)

        # зменяем ширину всех столбцов с оценками
        for i in range(3, int(sheets.max_column)):
            sheets.column_dimensions[get_column_letter(i)].width = 2

        # меняем тип ячейки на число, если это возможно
        for row in range(students_count-3):
            for col in range(3, int(sheets.max_column)):
                ch_range = get_column_letter(col) + str(row+3)
                if sheets[ch_range].value in marks_str:
                    sheets[ch_range].value = int(sheets[ch_range].value)

        i = 3
        for col in sheets.iter_rows(min_col=3, max_col=87, min_row=3, max_row=students_count-1):
            marks_count = 0
            n_count = 0
            for cell in col:
                if cell.value in marks_int:
                    marks_count += 1
                elif cell.value == 'н':
                    n_count += 1

            sheets['CJ'+str(i)].value = marks_count
            sheets['CK' + str(i)].value = n_count
            i += 1

        #######################################################
        ###########  Сравнение кол-ва уроков с УП  ############
        #######################################################

        up_class_index = uchebniy_plan[0].index(file.lower())

        up_lesson_index = 0
        for i in range (len(uchebniy_plan)):
            if uchebniy_plan[i][0] == lesson.lower():
                up_lesson_index = i
                #print(lesson.lower(), uchebniy_plan[up_lesson_index][up_class_index])
                break
        if up_lesson_index == 0:
            print('ПРОВЕРЬТЕ НАЗВАНИЕ ПРЕДМЕТА В УЧЕБНОМ ПЛАНЕ!!!\n'
                  'НАЗВАНИЕ ДОЛЖНО СОВПАДАТЬС ВЫГРУЗКОЙ ИЗ ЭЖД')
            sys.exit()

        up_lesson_nagruzka = uchebniy_plan[up_lesson_index][up_class_index]

        up_marks_number = up_lesson_nagruzka * 2 + 1
        if up_marks_number > 7:
            up_marks_number = 7

        #удаление пустых столбцов перед количеством оценок и пропусков
        i = 87
        flag = False
        while True:
            # Проверка, что столбец не содержит данных
            column_a = get_column_letter(i)

            for j in range (1,50):
                if sheets[column_a + str(j)].value not in ['', None]:
                    #print('Столбец ', column_a,' содержит данные')
                    flag = True
                    break
            if flag:
                break
            else:
                i -= 1

        m_range = 'CJ1:CK50'
        sheets.move_range(m_range, rows=0, cols=-(87-i))

        sheets.column_dimensions[get_column_letter(i + 1)].width = 11
        sheets.column_dimensions[get_column_letter(i + 2)].width = 11

        # очищаем формат листа
        for row in sheets.iter_rows():
            for cell in row:
                cell.border = None
                cell.font = None

        # ищем номер столбца с АП1
        for row in sheets.iter_rows():
            for cell in row:
                c = cell.value
                if c == 'АП1':
                    period_col = cell.column_letter

        # указываем название предмета в В2
        sheets['B2'] = lesson

        # проходим по фамилиям на листе журнала класса, не в файле средних баллов!!!
        for row in range(3, sheets.max_row+1):
            student_all = sheets['B' + str(row)].value

            try:
                student = ' '.join(student_all.split()[:2])
                student_mark_in_ap = int(sheets[period_col+str(row)].value)

                avg_row_student_id = students_indexes.index(student)
                avg_row_lesson_id = lessons_indexes.index(lesson)

                print('В АП1 у', avg_class_data[avg_row_student_id][0], 'стоит', avg_class_data[avg_row_student_id][avg_row_lesson_id+1])

                print('В средних баллах', lesson, ': у', student, '( id =', avg_row_student_id+7,'), mark_in_ap =', student_mark_in_ap)

                if avg_class_data[avg_row_student_id][avg_row_lesson_id+1] == student_mark_in_ap:
                #     avg_book_sheet
                # comment_class_book_cell = comment_class_book_sheet['G' + str(len(comment_class_book_sheet['A']))]
                # comment_class_book_cell.fill = PatternFill(fill_type='solid', fgColor="FFE697")
            except:
                continue






        avg_book.save('Средние баллы_' + file + '_ПРОВЕРЕН.xlsx')

    book.save('Промежуточный_' + file + '_ПРОВЕРЕНО.xlsx')
    book.close()







#
#         for row in range(students_count-3):
#             ch_cell = 'CJ' + str(row+7)
#             fill_cell = sheets[ch_cell]
#             #print(ch_cell, sheets[ch_cell].value)
#             if fill_cell.value >= up_marks_number:
#                 fill_cell.fill = PatternFill(fill_type='solid', fgColor="85EB6A")
#             else:
#                 fill_cell.fill = PatternFill(fill_type='solid', fgColor="FA7080")
#                 # добавляем замечние в файл КЛАССА
#                 comment_class_book_sheet.append([file, lesson, teacher, sheets['B' + str(row + 7)].value, up_marks_number,
#                                                  fill_cell.value, up_marks_number - fill_cell.value])
#
#                 if up_marks_number-fill_cell.value == 1:
#                     comment_class_book_cell = comment_class_book_sheet['G' + str(len(comment_class_book_sheet['A']))]
#                     comment_class_book_cell.fill = PatternFill(fill_type='solid', fgColor="FFE697")
#                 elif up_marks_number - fill_cell.value == 2:
#                     comment_class_book_cell = comment_class_book_sheet['G' + str(len(comment_class_book_sheet['A']))]
#                     comment_class_book_cell.fill = PatternFill(fill_type='solid', fgColor="FFD040")
#                 elif up_marks_number - fill_cell.value == 3:
#                     comment_class_book_cell = comment_class_book_sheet['G' + str(len(comment_class_book_sheet['A']))]
#                     comment_class_book_cell.fill = PatternFill(fill_type='solid', fgColor="F9A13E")
#
#                 # добавляем замечние в СВОДНЫЙ ОТЧЕТ
#                 comment_book_sheet.append(
#                     [file, lesson, teacher, sheets['B' + str(row + 7)].value, up_marks_number,
#                      fill_cell.value, up_marks_number - fill_cell.value])
#
#                 if up_marks_number-fill_cell.value == 1:
#                     comment_book_cell = comment_book_sheet['G' + str(len(comment_book_sheet['A']))]
#                     comment_book_cell.fill = PatternFill(fill_type='solid', fgColor="FFE697")
#                 elif up_marks_number - fill_cell.value == 2:
#                     comment_book_cell = comment_book_sheet['G' + str(len(comment_book_sheet['A']))]
#                     comment_book_cell.fill = PatternFill(fill_type='solid', fgColor="FFD040")
#                 elif up_marks_number - fill_cell.value == 3:
#                     comment_book_cell = comment_book_sheet['G' + str(len(comment_book_sheet['A']))]
#                     comment_book_cell.fill = PatternFill(fill_type='solid', fgColor="F9A13E")
#
#         #######################################################
#         #######################################################
#         #######################################################
#
#         # ставим дату и время проверки в ячейку B2
#         sheets['B4'] = date
#
#
#
#     # перемещаем обработанный файл в папку DONE
#     # source_file = 'data\\' + file + '.xlsx'
#     # destination_folder = 'done\\' + file + '.xlsx'
#     # shutil.move(source_file, destination_folder)
#
#     comment_class_book.save('comments\\' + file + ' ЗАМЕЧАНИЯ.xlsx')
#     comment_class_book.close()
#
#     comment_book.save('comments\ВСЕ ЗАМЕЧАНИЯ ПО ПРОВЕРКЕ НАКОПЛЯЕМОСТИ ОЦЕНОК.xlsx')
#     comment_book.close()