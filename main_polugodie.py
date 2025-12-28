# -*- coding: utf8 -*-
import asyncio
import shutil

from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter

from datetime import datetime
import os
import sys
import warnings

def find_data(sheet, search_value):
    for row in sheet.iter_rows():
        for cell in row:
            if cell.value == search_value:
                coordinate = cell.coordinate
                column = ""
                column_number = 0
                row_num = ""
                for char in coordinate:
                    if char.isalpha():  # Проверяем, является ли символ буквой
                        column += char
                    elif char.isdigit():  # Проверяем, является ли символ цифрой
                        row_num += char

                for i, letter in enumerate(reversed(column)):
                    column_number += (ord(letter.upper()) - ord('A') + 1) * (26 ** i)
                return column, column_number, row_num

    return None, None, None

warnings.simplefilter("ignore")

data_str = ['1', '2', '3', '4', '5']
marks_str = ['2', '3', '4', '5']
marks_int = [2, 3, 4, 5]

start_time = datetime.now()

# параллели для проверки
# grade_list = ['10','11']
grade_list = ['test']
data_date = "171225"

for grade in grade_list:
    # папка с журналами
    journals_folder = "journals_" + data_date + "\\" + grade
    # journals_folder = "journals\\"

    classes_korp_8 = ["5-А-З", "5-Б-З", "5-В-З", "5-Д-З", "5-Е-З", "5-З-З", "5-И-З", "5-Л-З", "5-Ф", "5-Ю", "5-Я",
                      "6-А-З", "6-Б-З", "6-В-З", "6-Д-З", "6-Е-З", "6-З-З", "6-Ю", "6-Я",
                      "7-А-З", "7-Б-З", "7-В-З", "7-Д-З", "7-Ц", "7-Ч", "7-Ш", "7-Э", "7-Ю", "7-Я",
                      "8-А-З", "8-Б-З", "8-В-З", "8-Д-З", "8-Ц", "8-Ч", "8-Ш", "8-Э", "8-Ю", "8-Я",
                      "9-А-З", "9-Б-З", "9-Ц", "9-Ч", "9-Ш", "9-Э", "9-Ю", "9-Я",
                      "10-Ф", "10-Ц", "10-Ч", "10-Ш", "10-Э", "10-Я",
                      "11-У", "11-Ц", "11-Ч", "11-Ш"]

    # Получаем текущую директорию
    current_dir = os.getcwd()

    # Создаем список для хранения имен файлов
    file_list = []

    # Перебираем файлы в текущей директории
    for root, dirs, files in os.walk(current_dir + '\\' + journals_folder):
        for filename in files:
            # for file_name_part in classes_korp_8:   # фильтр классов по корпусу 8
            #     if file_name_part not in filename:  # фильтр классов по корпусу 8
            #         continue                        # фильтр классов по корпусу 8
            if filename[-5:] == '.xlsx':
                file_list.append(filename)

    # книга замечаний по ОТМЕТКАМ
    log_book = Workbook()
    log_book_sheet = log_book.active
    log_book_sheet.append(['имя файла'])
    log_book_sheet.column_dimensions['A'].width = 100

    # книга замечаний по ОТМЕТКАМ
    comment_book_trimestr = Workbook()
    comment_book_trimestr_sheet = comment_book_trimestr.active
    comment_book_trimestr_sheet.append(['Класс',
                                        'Ученик',
                                        'Предмет',
                                        'Кол-во отметок',
                                        'Средний балл',
                                        'Выставленная отметка',
                                        'Правильная отметка',
                                        'Кол-во уроков',
                                        'Кол-во пропусков',
                                        'Процент пропусков',
                                        'Комментарии'])
    comment_book_trimestr_sheet.column_dimensions['A'].width = 6
    comment_book_trimestr_sheet.column_dimensions['B'].width = 50
    comment_book_trimestr_sheet.column_dimensions['C'].width = 50
    comment_book_trimestr_sheet.column_dimensions['D'].width = 5
    comment_book_trimestr_sheet.column_dimensions['E'].width = 5
    comment_book_trimestr_sheet.column_dimensions['F'].width = 5
    comment_book_trimestr_sheet.column_dimensions['G'].width = 5
    comment_book_trimestr_sheet.column_dimensions['H'].width = 5
    comment_book_trimestr_sheet.column_dimensions['I'].width = 5
    comment_book_trimestr_sheet.column_dimensions['J'].width = 5
    comment_book_trimestr_sheet.column_dimensions['K'].width = 25

    first_row = comment_book_trimestr_sheet["1"]
    for cell in first_row:
        style = Alignment(text_rotation=90, horizontal='center', vertical='center')
        cell.alignment = style

    # выбирается файл для обработки
    for file in file_list:
        file_name = file.split('.xlsx')[0]
        print(file_name, " | ", end="")

        journal_data = file_name.split(";")

        print(len(journal_data))

        if len(journal_data) != 3:
            log_book_sheet.append([file])
            log_book.save(r'comments\LOG ' + grade + ' ' + data_date + '.xlsx')
            continue

        class_name = journal_data[0]
        lesson = journal_data[1]
        group = journal_data[2]

        #######################################################################################
        ####################  Проверка журналов на выставление отметок  #######################
        #######################################################################################

        try:
            #############################################################################
            ###################### преобразование входных данных ########################
            #############################################################################

            # книга для проверки отметок
            book = load_workbook(journals_folder + "\\" + file)
            sheet = book.active

            # удаляем столбцы с ДЗ
            sheet.delete_cols(20, 23)

            # получаем список объединенных диапазонов и разъединяем объединенные ячейки в книге ОТМЕТОК
            merged_cells = list(map(str, sheet.merged_cells.ranges))

            for item in merged_cells:
                try:
                    sheet.unmerge_cells(item)
                except (KeyError, ValueError):
                    continue

            # перемещаем куски оценок в одну общую часть
            for part in range(1, 15):
                try:
                    m_range = 'C' + str(1 + part * 50) + ':S' + str((part + 1) * 50)
                    sheet.move_range(m_range, rows=-50 * part, cols=17 * part)
                except (ValueError, IndexError):
                    continue

            # считаем количество учеников на листе
            start_row = 3
            students_count = 1

            if sheet['A' + str(students_count)].value == "Нет оценок":
                book.close()
                continue

            while (sheet['A' + str(students_count + start_row)].value != '' and
                   sheet['A' + str(students_count + start_row)].value is not None):
                students_count += 1
                if students_count > 50:
                    break

            students_count -= 1

            # удаляем пустые строки
            if students_count < sheet.max_row:
                sheet.delete_rows(students_count, sheet.max_row - students_count + 1)

            # заменяем ширину всех столбцов с оценками
            for i in range(3, min(sheet.max_column, 100)):  # ограничиваем максимальное количество столбцов
                sheet.column_dimensions[get_column_letter(i)].width = 2

            # меняем тип ячейки на число, если это возможно
            for row in range(1, min(students_count - 2, 100)):  # ограничиваем количество строк
                for col in range(3, min(sheet.max_column, 250)):
                    ch_range = get_column_letter(col) + str(row + 3)
                    if sheet[ch_range].value in data_str:
                        sheet[ch_range].value = int(sheet[ch_range].value)

            # подсчет проведенных уроков на данную дату и удаление не нужных столбцов
            lessons_count = 0
            month_count = 1
            old_day_value = 0
            current_month = 9
            current_day = 0
            last_column_index = 1

            now_month = datetime.now().month
            now_day = datetime.now().day

            # считаем количество используемых столбцов
            max_col_int = 3
            while sheet[get_column_letter(max_col_int)+'3'].value in ['оц', 'вс', 'фк']:
                max_col_int += 1

            # заполняем пустые ячейки месяцами
            current_month = sheet[get_column_letter(col) + '1'].value
            for col in range(3, max_col_int):
                if sheet[get_column_letter(col) + '1'].value == None:
                    sheet[get_column_letter(col) + '1'].value = current_month
                else:
                    current_month = sheet[get_column_letter(col) + '1'].value

            # заполняем пустые ячейки днями
            current_day = sheet[get_column_letter(col) + '2'].value
            for col in range(3, max_col_int):
                if sheet[get_column_letter(col) + '2'].value == None:
                    sheet[get_column_letter(col) + '2'].value = current_day
                else:
                    current_day = sheet[get_column_letter(col) + '2'].value

            # Заменяет названия месяцев на их номера в заголовке
            replacements = {
                'янв': '1',
                'фев': '2',
                'мар': '3',
                'апр': '4',
                'май': '5',
                'сен': '9',
                'окт': '10',
                'ноя': '11',
                'дек': '12',
                'я': '1',
                'ф': '2',
                'а': '4',
                'с': '9',
                'о': '10',
                'н': '11',
                'д': '12'
            }

            for cell in sheet[1]:  # Первая строка
                if cell.value and isinstance(cell.value, str):
                    for old, new in replacements.items():
                        if old in cell.value:
                            cell.value = cell.value.replace(old, new)

            # считаем количество проведенных уроков
            lessons_list = []
            for col in range(3, max_col_int):
                # не учитываем продленные КТП и даты из 1 полугодия
                if sheet[get_column_letter(col) + '1'].value == '12':
                    break
                lessons_list.append(str(sheet[get_column_letter(col) + '2'].value) + '.' + str(sheet[get_column_letter(col) + '1'].value))

            lessons_count = len(set(lessons_list))

            #############################################################################
            ################### обработка преобразованных журналов ######################
            #############################################################################

            # поиск П1 на листе
            found = False
            t1_column = None
            t1_column_int = None

            for row in sheet.iter_rows():
                for cell in row:
                    # П1 проверка латиницы и кириллицы
                    if cell.value == 'П1':
                        t1_column_int = cell.column
                        t1_column = get_column_letter(cell.column)
                        found = True
                        break
                if found:
                    break

            if not found:
                # print('Не выставлен ПОЛУГОДИЕ 1')
                t1_column_int = sheet.max_column

            # поиск столбцов с оценкой и коэффициентом
            marks_columns_on_sheet = []
            coefficient_columns_on_sheet = []

            for row in sheet.iter_rows():
                for cell in row:
                    if cell.value == 'оц' and cell.offset(0, 1).value == 'вс':
                        if cell.column < t1_column_int:
                            marks_columns_on_sheet.append(cell.column)
                            coefficient_columns_on_sheet.append(cell.column + 1)

            # подсчет количества оценок по ученикам
            for row in range(4, min(students_count, 100)):  # ограничиваем количество строк
                if row > sheet.max_row:
                    break

                # начало обработки журнала по каждому ученику (построчно)
                marks_count_for_student = 0
                marks_sum_for_student = 0
                current_student_cell = sheet['B' + str(row)]
                if current_student_cell.value is None:
                    continue

                comment_for_student = []
                marks_count_clear = 0
                propuski_count = 0

                current_student = str(current_student_cell.value)[:-2] if current_student_cell.value else ""

                for i in range(len(marks_columns_on_sheet)):
                    try:

                        mark_cell = sheet[get_column_letter(marks_columns_on_sheet[i]) + str(row)]
                        coeff_cell = sheet[get_column_letter(coefficient_columns_on_sheet[i]) + str(row)]

                        # пропускаем ячейки без отметок и с "н"
                        if mark_cell.value == "": continue
                        elif mark_cell.value == "н": propuski_count += 1

                        if mark_cell.value in marks_int and coeff_cell.value is not None:
                            try:
                                marks_count_clear += 1
                                marks_count_for_student += int(coeff_cell.value)
                                marks_sum_for_student += int(mark_cell.value) * int(coeff_cell.value)
                            except (ValueError, TypeError):
                                print(f"Ошибка в обработке оценок: {e}")
                                continue
                    except (ValueError, IndexError, KeyError) as e:
                        print(f"Ошибка в обработке оценок: {e}")
                        continue

                # проверка деления на 0
                try:
                    avg_mark_for_student = round(marks_sum_for_student / marks_count_for_student,
                                                 2) if marks_count_for_student > 0 else 0
                except ZeroDivisionError:
                    avg_mark_for_student = 0

                # проверяем на НПА и преобразуем среднее в оценку по правилу < .6
                avg_mark_for_student_ejd = 'НПА'
                if marks_count_for_student == 0:
                    comment_for_student.append('Нет отметок')
                elif 2 < lessons_count <= 17 and marks_count_clear < 3:
                    comment_for_student.append('Не хватает отметок: ' + str(3 - marks_count_clear))
                elif 17 < lessons_count <= 35 and marks_count_clear < 5:
                    comment_for_student.append('Не хватает отметок: ' + str(5 - marks_count_clear))
                elif lessons_count > 35 and marks_count_clear < 7:
                    comment_for_student.append('Не хватает отметок: ' + str(7 - marks_count_clear))
                else:
                    if avg_mark_for_student < 2.6:
                        avg_mark_for_student_ejd = 'А/З'
                        comment_for_student.append('Выходит А/З')
                    elif avg_mark_for_student < 3.6:
                        avg_mark_for_student_ejd = 3
                    elif avg_mark_for_student < 4.6:
                        avg_mark_for_student_ejd = 4
                    else:
                        avg_mark_for_student_ejd = 5

                # отметка ученика за 1 полугодие
                if t1_column is not None and str(t1_column).strip() != '':
                    trimestr_cell_value = sheet[t1_column + str(row)].value

                    if trimestr_cell_value is None or trimestr_cell_value == '':
                        comment_for_student.append('Нет отметки за полугодие')
                    else:
                        # print(trimestr_cell_value)

                        if str(trimestr_cell_value) != str(avg_mark_for_student_ejd):
                            if str(trimestr_cell_value) == 'Зч':
                                continue
                            elif str(avg_mark_for_student_ejd) == 'НПА':
                                continue
                            # print(current_student, trimestr_cell_value, avg_mark_for_student_ejd)
                            comment_for_student.append('Выставлена неправильная отметка: ' + str(trimestr_cell_value))
                else:
                    trimestr_cell_value = ''
                    comment_for_student.append('Нет отметки за полугодие')

                propuski_percent = round(propuski_count/lessons_count*100, 0)

                if len(comment_for_student) > 0:
                    comment_book_trimestr_sheet.append([class_name,
                                                    current_student,
                                                    lesson,
                                                    marks_count_clear,
                                                    avg_mark_for_student,
                                                    trimestr_cell_value,
                                                    avg_mark_for_student_ejd,
                                                    lessons_count,
                                                    propuski_count,
                                                    propuski_percent,
                                                    '\n'.join(comment_for_student)])

            comment_book_trimestr.save(
                r'comments\ВСЕ ЗАМЕЧАНИЯ ПО ПРОВЕРКЕ ПОЛУГОДИЯ 1 ' + grade + ' ' + data_date + '.xlsx')

            # book.save(file[:-5] + grade + ' ' + data_date + '.xlsx')


        except Exception as e:
            # Получаем информацию о последнем исключении
            exc_type, exc_value, exc_traceback = sys.exc_info()
            # Получаем номер строки кода, где произошла ошибка
            line_number = exc_traceback.tb_lineno
            print(f"Ошибка при обработке файла {file} (строка кода {line_number}): {e}")
            log_book_sheet.append([file + f" - ОШИБКА в строке кода {line_number}: " + str(e)])
            log_book.save(r'comments\LOG ' + grade + ' ' + data_date + '.xlsx')

    # Закрываем книги после завершения всех операций
    try:
        comment_book_trimestr.close()
        log_book.close()
    except:
        pass

end_time = datetime.now()

print()
print(start_time, end_time)